"""Directive 54: ERCOT Interconnection Queue Infiltration.

Scrapes the ERCOT Generator Interconnection & Change Request (GICR) Queue
to identify where 500MW+ requests are clustering in Dallas and Austin.

Target: Land OUTSIDE these clusters to avoid 7-year transmission upgrade wait.

Usage:
    python data_center/ercot_queue.py
"""

import sys, os, csv, re, json, sqlite3, io
from pathlib import Path
from datetime import datetime

try:
    import requests
except ImportError:
    requests = None

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = str(Path(SCRIPT_DIR).parent / "deals.db")

# ERCOT queue sources
ERCOT_QUEUE_URLS = [
    "https://www.ercot.com/mp/data-products/data-product-details?id=NP4-201-ER",
    "https://www.ercot.com/mp/data-products/data-product-details?id=NP4-200-ER",
    "https://www.ercot.com/mp/data-products/data-product-details?id=NP3-560-ER",
]

# Alternative direct CSV endpoints (historical patterns)
ERCOT_CSV_CANDIDATES = [
    "https://www.ercot.com/content/cdr/html/GICR",
    "https://www.ercot.com/misapp/servlets/IceDocServlet?docGroup=GICR",
    "https://www.ercot.com/content/wcm/lists/226656/GICR_Queue.xlsx",
    "https://www.ercot.com/content/wcm/lists/226656/GICR.xlsx",
    "https://www.ercot.com/content/wcm/lists/226656/Generator_Interconnection_Status.xlsx",
]


def ensure_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ercot_queue (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT,
            county TEXT,
            state TEXT,
            requested_mw REAL,
            fuel_type TEXT,
            status TEXT,
            study_status TEXT,
            interconnection_customer TEXT,
            ercot_region TEXT,
            submitted_date TEXT,
            cod_date TEXT,
            lat REAL,
            lng REAL,
            cluster_zone TEXT,
            scraped_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ercot_clusters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cluster_name TEXT,
            county TEXT,
            total_mw REAL,
            project_count INTEGER,
            avg_mw REAL,
            min_lat REAL,
            max_lat REAL,
            min_lng REAL,
            max_lng REAL,
            region TEXT
        )
    """)
    conn.commit()


def scrape_ercot():
    """Attempt to scrape the ERCOT interconnection queue."""
    print("=" * 60)
    print("DIRECTIVE 54: ERCOT Queue Infiltration")
    print("=" * 60)

    conn = sqlite3.connect(DB_PATH)
    ensure_table(conn)

    # Check if already cached
    existing = conn.execute("SELECT COUNT(*) FROM ercot_queue").fetchone()[0]
    if existing > 0:
        print(f"ERCOT queue already cached: {existing} projects. Use --force to refresh.")

        # Show clusters
        clusters = conn.execute("""
            SELECT county, COUNT(*) as cnt, ROUND(SUM(requested_mw)) as total_mw,
                   ROUND(AVG(requested_mw),0) as avg_mw
            FROM ercot_queue
            WHERE requested_mw >= 500
              AND county IS NOT NULL
            GROUP BY county
            ORDER BY total_mw DESC
        """).fetchall()

        if clusters:
            print("\n  500MW+ clusters by county:")
            print(f"  {'County':<20} {'Count':>6} {'Total MW':>10} {'Avg MW':>8}")
            print(f"  {'-'*20} {'-'*6} {'-'*10} {'-'*8}")
            for c in clusters:
                print(f"  {c[0]:<20} {c[1]:>6} {c[2]:>10.0f} {c[3]:>8.0f}")

        conn.close()
        return

    if not requests:
        print("requests library required. pip install requests")
        return

    headers = {
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    # Try to find the actual queue file on the ERCOT data portal
    print("Scanning ERCOT data portal for queue file...")

    found_urls = []
    for url in ERCOT_QUEUE_URLS:
        try:
            r = requests.get(url, headers=headers, timeout=20)
            if r.status_code == 200:
                # Parse the page for download links
                for match in re.finditer(r'href=["\']([^"\']*\.(?:xlsx|csv))["\']', r.text, re.I):
                    href = match.group(1)
                    if href.startswith("/"):
                        href = "https://www.ercot.com" + href
                    found_urls.append(href)
                    print(f"  Found: {href}")
        except Exception as e:
            print(f"  Error fetching {url}: {e}")

    # Try the known CSV candidates
    print("\nTrying direct CSV candidates...")
    for url in ERCOT_CSV_CANDIDATES:
        try:
            r = requests.head(url, headers=headers, timeout=10, allow_redirects=True)
            ct = r.headers.get("Content-Type", "")
            cl = r.headers.get("Content-Length", "?")
            print(f"  {r.status_code}: {url} ({ct}, {cl}b)")
            if r.status_code == 200 and ("csv" in ct.lower() or "excel" in ct.lower() or "spreadsheet" in ct.lower()):
                found_urls.append(url)
        except Exception as e:
            print(f"  Error: {e}")

    # Download and parse first found file
    data = []
    for url in found_urls:
        print(f"\nDownloading: {url}")
        try:
            r = requests.get(url, headers=headers, timeout=60)
            if r.status_code != 200:
                continue

            content = r.content
            print(f"  Size: {len(content)} bytes")

            # Try parsing as CSV
            try:
                decoded = content.decode("utf-8", errors="replace")
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
                if rows:
                    print(f"  CSV rows: {len(rows)}, columns: {list(rows[0].keys())[:15]}")
                    data = rows
                    break
            except Exception:
                pass

            # Try Excel
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                headers_row = [cell.value for cell in next(ws.iter_rows())]
                print(f"  Excel sheet: {wb.sheetnames[0]}, columns: {headers_row[:15]}")
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(dict(zip(headers_row, row)))
                data = rows
                break
            except ImportError:
                print("  openpyxl not installed, cannot parse Excel")
            except Exception as e:
                print(f"  Excel parse error: {e}")

        except Exception as e:
            print(f"  Download error: {e}")

    if not data:
        print("\nCould not auto-download ERCOT queue. Creating manual entry template.")
        print("Download manually from: https://www.ercot.com/mp/data-products")
        print("Then run: python data_center/ercot_queue.py --import <filepath>")
        conn.close()
        return

    # Map known ERCOT columns
    col_map = {
        "Project Name": "project_name", "Project": "project_name",
        "County": "county", "County Name": "county",
        "State": "state",
        "Requested MW": "requested_mw", "MW": "requested_mw", "Capacity (MW)": "requested_mw",
        "Fuel Type": "fuel_type", "Fuel": "fuel_type",
        "Status": "status", "Current Status": "status",
        "Study Status": "study_status",
        "Interconnection Customer": "interconnection_customer", "Customer": "interconnection_customer",
        "ERCOT Region": "ercot_region", "Region": "ercot_region",
        "Submitted Date": "submitted_date", "Date Submitted": "submitted_date",
        "COD": "cod_date", "Commercial Operation Date": "cod_date",
    }

    mapped = 0
    for row in data:
        rec = {}
        for csv_col, db_col in col_map.items():
            val = row.get(csv_col, row.get(csv_col.lower(), ""))
            if val is not None:
                rec[db_col] = str(val).strip()

        if not rec.get("project_name"):
            continue

        try:
            conn.execute("""
                INSERT INTO ercot_queue
                (project_name, county, state, requested_mw, fuel_type, status,
                 study_status, interconnection_customer, ercot_region,
                 submitted_date, cod_date, scraped_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                rec.get("project_name", "")[:200],
                rec.get("county", "")[:100],
                rec.get("state", "")[:10],
                float(re.sub(r"[^\d.]", "", rec.get("requested_mw", "0"))) if rec.get("requested_mw") else None,
                rec.get("fuel_type", "")[:50],
                rec.get("status", "")[:50],
                rec.get("study_status", "")[:50],
                rec.get("interconnection_customer", "")[:200],
                rec.get("ercot_region", "")[:50],
                rec.get("submitted_date", "")[:20],
                rec.get("cod_date", "")[:20],
                datetime.utcnow().isoformat(),
            ))
            mapped += 1
        except Exception as e:
            print(f"  Insert error: {e}")

    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM ercot_queue").fetchone()[0]
    print(f"\nInserted {mapped} projects (total: {total})")

    # Analyze 500MW+ clusters
    print("\n--- 500MW+ Cluster Analysis ---")
    clusters = conn.execute("""
        SELECT
            CASE
                WHEN county LIKE 'DALLAS%' OR county LIKE 'TARRANT%' OR county LIKE 'COLLIN%'
                     OR county LIKE 'DENTON%' OR county LIKE 'ELLIS%' THEN 'DFW'
                WHEN county LIKE 'TRAVIS%' OR county LIKE 'WILLIAMSON%' OR county LIKE 'HAYS%'
                     THEN 'AUSTIN'
                WHEN county LIKE 'BEXAR%' OR county LIKE 'COMAL%' THEN 'SAN ANTONIO'
                WHEN county LIKE 'HARRIS%' OR county LIKE 'FORT BEND%' OR county LIKE 'MONTGOMERY%'
                     THEN 'HOUSTON'
                ELSE 'OTHER'
            END as region,
            county,
            COUNT(*) as project_count,
            ROUND(SUM(requested_mw)) as total_mw,
            ROUND(AVG(requested_mw), 0) as avg_mw,
            ROUND(MIN(requested_mw), 0) as min_mw,
            ROUND(MAX(requested_mw), 0) as max_mw
        FROM ercot_queue
        WHERE requested_mw >= 500
        GROUP BY region, county
        ORDER BY total_mw DESC
    """).fetchall()

    if clusters:
        print(f"\n  {'Region':<16} {'County':<20} {'Count':>5} {'MW':>10} {'Avg':>8}")
        print(f"  {'-'*16} {'-'*20} {'-'*5} {'-'*10} {'-'*8}")
        for c in clusters:
            print(f"  {c[0]:<16} {c[1]:<20} {c[2]:>5} {c[3]:>10.0f} {c[4]:>8.0f}")

        # Write cluster recommendations
        conn.execute("DELETE FROM ercot_clusters")

        region_data = {}
        for c in clusters:
            region = c[0]
            if region not in region_data:
                region_data[region] = {"count": 0, "total_mw": 0, "counties": set()}
            region_data[region]["count"] += c[2]
            region_data[region]["total_mw"] += c[3]
            region_data[region]["counties"].add(c[1])

        for region, rd in region_data.items():
            conn.execute("""
                INSERT INTO ercot_clusters
                (cluster_name, county, total_mw, project_count, avg_mw, region)
                VALUES (?,?,?,?,?,?)
            """, (
                f"{region} Interconnection Cluster",
                ", ".join(sorted(rd["counties"])),
                rd["total_mw"],
                rd["count"],
                rd["total_mw"] / rd["count"] if rd["count"] else 0,
                region,
            ))

        conn.commit()
        print("\nCluster zones written to ercot_clusters.")

        # Land targeting insight
        print("\n--- Land Strategy ---")
        for region in ["DFW", "AUSTIN", "HOUSTON", "SAN ANTONIO"]:
            rd = region_data.get(region)
            if rd:
                print(f"  {region}: {rd['total_mw']:,.0f}MW queued")
                print(f"    → Target land OUTSIDE {region} core, within {rd['count']}mi of 345kV")
        print("    → Avoid 7-year transmission upgrade wait by choosing exurban sites")
        print("    → Prioritize ONCOR-served land NOT in high-queue counties")
    else:
        print("No 500MW+ projects found. Data may need manual import.")

    conn.close()
    return total


def manual_import(filepath):
    """Import a manually downloaded ERCOT CSV/Excel file."""
    conn = sqlite3.connect(DB_PATH)
    ensure_table(conn)
    # same logic as above but reading from filepath
    print(f"Manual import from {filepath} not yet implemented — extend this function.")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="ERCOT Interconnection Queue scraper")
    parser.add_argument("--force", action="store_true", help="Re-scrape even if cached")
    parser.add_argument("--import", dest="import_file", help="Import from local file")
    args = parser.parse_args()

    if args.import_file:
        manual_import(args.import_file)
    else:
        scrape_ercot()
